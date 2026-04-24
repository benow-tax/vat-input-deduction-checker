# -*- coding: utf-8 -*-
"""
㈜비나우 매입세액불공제 자동 검출 - Streamlit 웹앱
"""

import streamlit as st
import pandas as pd
import re, math, io, json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 페이지 설정 ───────────────────────────────────────
st.set_page_config(
    page_title='비나우 매입불공제 자동 검출',
    page_icon='🔍',
    layout='wide',
    initial_sidebar_state='expanded',
)

# ── 커스텀 CSS ────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');

html, body, [class*="css"] { font-family: 'Noto Sans KR', sans-serif; }

.main-header {
    background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
    color: white;
    padding: 2rem 2.5rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
}
.main-header h1 { font-size: 1.8rem; font-weight: 700; margin: 0 0 0.3rem 0; }
.main-header p  { font-size: 0.9rem; opacity: 0.85; margin: 0; }

.guide-box {
    background: #F0F7FF;
    border-left: 4px solid #2E75B6;
    border-radius: 0 8px 8px 0;
    padding: 1rem 1.2rem;
    margin-bottom: 0.8rem;
}
.guide-box h4 { color: #1F3864; margin: 0 0 0.4rem 0; font-size: 0.95rem; }
.guide-box p, .guide-box li { color: #333; font-size: 0.85rem; line-height: 1.6; margin: 0; }

.result-card {
    border-radius: 10px;
    padding: 1rem 1.2rem;
    margin-bottom: 0.6rem;
}
.card-red    { background: #FFF0F0; border-left: 4px solid #C00000; }
.card-yellow { background: #FFFBF0; border-left: 4px solid #BF8F00; }
.card-blue   { background: #F0F4FF; border-left: 4px solid #1F4E79; }
.card-gray   { background: #F5F5F5; border-left: 4px solid #595959; }

.stat-box {
    text-align: center;
    padding: 1rem;
    border-radius: 10px;
    background: white;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}
.stat-box .num  { font-size: 2rem; font-weight: 700; }
.stat-box .lbl  { font-size: 0.8rem; color: #666; margin-top: 0.2rem; }
.num-red    { color: #C00000; }
.num-yellow { color: #BF8F00; }
.num-blue   { color: #1F4E79; }
.num-gray   { color: #595959; }

.stTabs [data-baseweb="tab-list"] { gap: 8px; }
.stTabs [data-baseweb="tab"] {
    background: #F0F4F8;
    border-radius: 8px 8px 0 0;
    padding: 0.5rem 1.2rem;
    font-weight: 600;
}
.stTabs [aria-selected="true"] {
    background: #1F3864 !important;
    color: white !important;
}

.step-badge {
    display: inline-block;
    background: #1F3864;
    color: white;
    border-radius: 50%;
    width: 24px;
    height: 24px;
    text-align: center;
    line-height: 24px;
    font-size: 0.8rem;
    font-weight: 700;
    margin-right: 8px;
}
</style>
""", unsafe_allow_html=True)

# ── 기본 설정값 ───────────────────────────────────────
DEFAULT_CONFIG = {
    'taxi_keywords': ['택시','대리운전','콜밴','티머니택시','카카오T.*택시','카카오택시','onda','티머니onda','항공권','항공료','항공.*요금','대한항공','아시아나','제주항공','진에어','티웨이','에어서울','에어부산','이스타항공'],
    'entertain_keywords': ['인플루언서.*식사','인플루언서.*접대','인플루언서.*미팅','거래처.*접대','거래처.*미팅','거래처.*식사','광고모델.*음료','광고모델.*접대','모델.*음료대','미팅.*음료 접대','본사 미팅.*접대','접대','주차권','사무실.*방문.*주차','방문.*주차비','NCT.*음료','나연.*음료'],
    'store_keywords': ['아지트 뉴욕','아지트뉴욕','미국 매장','미국매장','일본 매장','일본매장','일본 아지트','일본아지트'],
    'us_amazon_keywords': ['미국.*아마존','아마존.*미국','비나우뷰티.*아마존','아마존.*비나우뷰티'],
    'exempt_cars': ['216다 9109','216다9109'],
    'store_exclude_keywords': ['메가와리'],
    'note_general': '세부 계정과목 비용인지 선급금인지 확인 후 불공제 여부 결정하기',
    'note_us_amazon': '본사, 비나우뷰티 중 어디 귀속 비용인지 확인하기 / 비나우뷰티 귀속 비용이라면 세부 계정과목 비용인지 선급금인지 확인 후 불공제 여부 결정하기',
}

if 'config' not in st.session_state:
    st.session_state.config = DEFAULT_CONFIG.copy()

# ── 판단 로직 ─────────────────────────────────────────
AVIATION_FREIGHT_EXEMPT  = ['운송','운임','화물','FBA','물류','항공 운송','항공운송']
CAR_KEYWORDS             = ['법인차','법인차량','렌트료','리스료','운전대행료']
POST_KEYWORDS            = ['우체국','우정사업본부']
SATAEK_KEYWORDS          = ['사택']
BRANCH_KEYWORDS          = ['일본지사','미국지사','지사 파견','지사파견','해외 파견','해외파견']
IMPORT_TAX_KEYWORDS      = ['수입관세']
GOLD_KEYWORDS            = ['금 매입','금매입']
INICIS_KEYWORDS          = ['이니시스','INICIS','inicis']
DOCUMENT_KEYWORDS        = ['서류','발급','CFS','증명','cfs']
HR_DEPT_KEYWORDS         = ['HR','채용','본부장']
MARKETING_DEPT_KEYWORDS  = ['마케팅','marketing','MD','브랜드','콘텐츠']
SUBSIDIARY_EXPORT_EXEMPT = ['BL발급','핸들링','FOB','수출 핸들링','수출핸들링','수출면장','FBA','항공 운송','항공운송','해상 운송','해상운송']

def contains_any(text, patterns, regex=True):
    if not isinstance(text, str): return False
    for p in patterns:
        try:
            if regex:
                if re.search(p, text, re.IGNORECASE): return True
            else:
                if p.lower() in text.lower(): return True
        except:
            if p.lower() in text.lower(): return True
    return False

def is_exempt_car(text, exempt_cars):
    for car in exempt_cars:
        if car.replace(' ','') in text.replace(' ',''): return True
    return '레이' in text

def classify_row(row, cfg):
    desc=str(row.get('적요','') or ''); vendor=str(row.get('거래처','') or '')
    supply=float(row.get('공급가액',0) or 0); vat=float(row.get('세액',0) or 0); total=float(row.get('합계',0) or 0)
    dept=str(row.get('비용센터','') or '')+str(row.get('작성부서','') or '')
    NG=cfg.get('note_general', DEFAULT_CONFIG['note_general'])
    NA=cfg.get('note_us_amazon', DEFAULT_CONFIG['note_us_amazon'])

    if contains_any(desc,cfg['taxi_keywords']) or contains_any(vendor,cfg['taxi_keywords']):
        if not contains_any(desc,AVIATION_FREIGHT_EXEMPT):
            return ('불공제','','택시·대리운전·여객항공: 여객운송업자로부터의 매입, 세금계산서 발급 불가')

    if contains_any(desc,CAR_KEYWORDS) or contains_any(vendor,CAR_KEYWORDS):
        if is_exempt_car(desc,cfg['exempt_cars']) or is_exempt_car(vendor,cfg['exempt_cars']): return None
        if not bool(re.search(r'\d{2,3}[가-힣]{1,2}\s?\d{4}',desc)) and '레이' not in desc:
            return ('담당자확인','','법인차량 관련 비용이나 차량번호 미기재 → 레이(경차)이면 공제, 그 외 법인차이면 불공제')
        return ('불공제','비영업용소형승용차구입, 유지 및 임차','법인차량(비경차) 관련 비용')

    if contains_any(desc,POST_KEYWORDS) or contains_any(vendor,POST_KEYWORDS):
        if vat==0 and supply==total: return ('불공제','일반면세','우체국: 세액=0, 면세거래(직접방문 접수)')
        return None

    if contains_any(desc,SATAEK_KEYWORDS):
        if contains_any(desc,BRANCH_KEYWORDS):
            return ('불공제','사업과 관련없는 지출','지사 파견 인원 또는 해외지사 직원 사택 관련 비용 (임대료·물품 포함)')
        return None

    if contains_any(desc,IMPORT_TAX_KEYWORDS): return ('불공제','사업과 관련없는 지출','수입관세: 납부 추적 불가로 보수적 불공제 처리')
    if contains_any(desc,GOLD_KEYWORDS): return ('불공제','사업과 관련없는 지출','금 매입 관련')
    if contains_any(vendor,INICIS_KEYWORDS) and contains_any(desc,DOCUMENT_KEYWORDS):
        return ('불공제','사업과 관련없는 지출','이니시스를 통한 서류발급(CFS 등): 세금계산서 발급 불가 구조')

    if not (contains_any(desc,SUBSIDIARY_EXPORT_EXEMPT) and '선급금' not in desc):
        if contains_any(desc,cfg['us_amazon_keywords']): return ('담당자확인','',NA)
        excl=cfg.get('store_exclude_keywords',[])
        if contains_any(desc,cfg['store_keywords']) and not any(k in desc for k in excl):
            return ('담당자확인','',NG)

    if '커피챗' in desc:
        if contains_any(dept,MARKETING_DEPT_KEYWORDS): return None
        if contains_any(dept,HR_DEPT_KEYWORDS):
            if contains_any(desc,['음료','카페','커피','음식','식사','디저트','간식']):
                return ('불공제','접대비관련매입세액','면접(커피챗): 면접자 음료·음식 제공 → 접대비')
            return ('담당자확인','','커피챗 항목: 비용센터가 HR/채용/본부장 → 면접 여부 확인 필요')
        if '사내행사' in desc: return None
        return ('담당자확인','','커피챗 항목: 비용센터 불명 → 면접(HR/본부장)인지 행사(마케팅)인지 확인 필요')

    if contains_any(desc,cfg['entertain_keywords']):
        return ('불공제','접대비관련매입세액','외부인(거래처·인플루언서·광고모델 등) 접대 비용')
    return None

def load_data(file):
    name = file.name.lower()
    if name.endswith('.csv'):
        for enc in ['utf-8','cp949','euc-kr']:
            try: return pd.read_csv(file, encoding=enc)
            except: file.seek(0)
    return pd.read_excel(file, header=0)

def run_check(df, cfg):
    for col in ['공급가액','세액','합계']:
        if col in df.columns:
            df[col]=pd.to_numeric(df[col].astype(str).str.replace(',','').str.strip(),errors='coerce').fillna(0)
    if '(세금)계산서일' in df.columns:
        df['_is_date']=pd.to_datetime(df['(세금)계산서일'],errors='coerce').notna()
    else:
        df['_is_date']=True
    actual=df[df['_is_date']].copy()
    results={'불공제_조정필요':[],'담당자_확인필요':[],'세무구분_오류':[],'적요_공란':[]}
    for _,row in actual.iterrows():
        tax=str(row.get('세무','') or ''); vat=float(row.get('세액',0) or 0)
        sup=float(row.get('공급가액',0) or 0); tot=float(row.get('합계',0) or 0)
        desc=str(row.get('적요','') or '').strip()
        done=tax in ['매입불공제','면세매입','영세매입','현금영수증매입','수입','카드(관리용/불공)']
        if not desc or desc=='nan':
            results['적요_공란'].append(row.to_dict())
        if tax in ['과세매입','카드매입'] and vat==0 and sup==tot:
            d=row.to_dict(); d['_오류유형']='[오류A] 세액=0인데 과세매입/카드매입'; d['_비고']='세액=0이면 영세매입·면세매입·매입불공제 중 하나로 변경 필요'
            results['세무구분_오류'].append(d); continue
        if tax in ['영세매입','면세매입'] and vat>0:
            d=row.to_dict(); d['_오류유형']='[오류B] 세액>0인데 영세매입/면세매입'; d['_비고']='세액이 있으면 과세매입 등으로 변경 필요'
            results['세무구분_오류'].append(d); continue
        if done or tax not in ['과세매입','카드매입']: continue
        res=classify_row(row,cfg)
        if res is None: continue
        cat,reason,note=res
        d=row.to_dict(); d['_불공제사유']=reason; d['_비고']=note
        if cat=='불공제': results['불공제_조정필요'].append(d)
        elif cat=='담당자확인': results['담당자_확인필요'].append(d)
    return results

# ── 엑셀 출력 ─────────────────────────────────────────
COLOR={'red_header':'C00000','red_light':'FFCCCC','yellow_header':'BF8F00','yellow_light':'FFF2CC','blue_header':'1F4E79','blue_light':'DEEAF1','gray_header':'595959','gray_light':'EDEDED','white':'FFFFFF','summary_bg':'F2F2F2'}
SECTION_CONFIG={
    '불공제_조정필요':{'header_color':COLOR['red_header'],'row_color':COLOR['red_light'],'label':'🔴 불공제 조정 필요'},
    '담당자_확인필요':{'header_color':COLOR['yellow_header'],'row_color':COLOR['yellow_light'],'label':'🟡 담당자 확인 필요'},
    '세무구분_오류':{'header_color':COLOR['blue_header'],'row_color':COLOR['blue_light'],'label':'⚠️ 세무구분 오류'},
    '적요_공란':{'header_color':COLOR['gray_header'],'row_color':COLOR['gray_light'],'label':'📋 적요 공란'},
}
BASE_COLS=['(세금)계산서일','적요','세무','사유','거래처','거래처사업자번호','공급가액','세액','합계','전표번호','작성부서','비용센터']
RESULT_EXTRA={'불공제_조정필요':['_불공제사유','_비고'],'담당자_확인필요':['_비고'],'세무구분_오류':['_오류유형','_비고'],'적요_공란':['_비고']}
COL_HEADERS={'(세금)계산서일':'계산서일','적요':'적요','세무':'세무구분','사유':'기존사유','거래처':'거래처','거래처사업자번호':'사업자번호','공급가액':'공급가액','세액':'세액','합계':'합계','전표번호':'전표번호','작성부서':'작성부서','비용센터':'비용센터','_불공제사유':'▶ 불공제사유(ERP입력용)','_비고':'비고','_오류유형':'오류유형'}
COL_WIDTHS={'계산서일':14,'적요':60,'세무구분':14,'기존사유':22,'거래처':22,'사업자번호':16,'공급가액':14,'세액':12,'합계':14,'전표번호':20,'작성부서':14,'비용센터':14,'▶ 불공제사유(ERP입력용)':30,'비고':40,'오류유형':30}

def mf(c): return PatternFill('solid',fgColor=c)
def mb(): s=Side(style='thin',color='BFBFBF'); return Border(left=s,right=s,top=s,bottom=s)

def write_section(ws,results,sk,sr):
    sc=SECTION_CONFIG[sk]; data=results[sk]; cols=BASE_COLS+RESULT_EXTRA[sk]
    c=ws.cell(row=sr,column=1,value=f"{sc['label']}  ({len(data)}건)")
    c.font=Font(bold=True,size=12,color='FFFFFF'); c.fill=mf(sc['header_color'])
    ws.merge_cells(start_row=sr,start_column=1,end_row=sr,end_column=len(cols)); sr+=1
    if not data: ws.cell(row=sr,column=1,value='해당 없음').font=Font(italic=True,color='888888'); return sr+2
    hf=mf(sc['header_color'])
    for ci,col in enumerate(cols,1):
        c=ws.cell(row=sr,column=ci,value=COL_HEADERS.get(col,col)); c.fill=hf
        c.font=Font(bold=True,color='FFFFFF',size=10); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=mb()
    sr+=1; rf=mf(sc['row_color']); wf=mf(COLOR['white'])
    for ri,rec in enumerate(data):
        fill=rf if ri%2==0 else wf
        for ci,col in enumerate(cols,1):
            val=rec.get(col,'')
            if isinstance(val,float):
                try: val=('' if math.isnan(val) else (int(val) if val==int(val) else val))
                except: pass
            c=ws.cell(row=sr,column=ci,value=val); c.fill=fill; c.border=mb()
            c.alignment=Alignment(vertical='center',wrap_text=(col in ['적요','_비고','_오류유형','_불공제사유']))
            if col in ['공급가액','세액','합계']: c.number_format='#,##0'; c.alignment=Alignment(horizontal='right',vertical='center')
        sr+=1
    return sr+2

def make_excel(results, filename):
    wb=Workbook(); ws=wb.active; ws.title='검토요약'
    ws.column_dimensions['A'].width=30; ws.column_dimensions['B'].width=15; ws.column_dimensions['C'].width=20
    c=ws.cell(row=1,column=1,value='㈜비나우 매입세액불공제 검토 결과 요약'); c.font=Font(bold=True,size=14,color='FFFFFF'); c.fill=mf('1F4E79'); ws.merge_cells('A1:C1')
    ws.cell(row=2,column=1,value=f'검토 파일: {filename}').font=Font(size=10,color='595959')
    ws.cell(row=3,column=1,value=f'검토 일시: {datetime.now().strftime("%Y-%m-%d %H:%M")}').font=Font(size=10,color='595959')
    for ci,h in enumerate(['구분','건수','세액 합계'],1):
        c=ws.cell(row=5,column=ci,value=h); c.fill=mf('2E75B6'); c.font=Font(bold=True,color='FFFFFF'); c.alignment=Alignment(horizontal='center'); c.border=mb()
    cm={'불공제_조정필요':COLOR['red_light'],'담당자_확인필요':COLOR['yellow_light'],'세무구분_오류':COLOR['blue_light'],'적요_공란':COLOR['gray_light']}
    tv=0
    for ri,(key,sc) in enumerate(SECTION_CONFIG.items(),6):
        data=results[key]; cnt=len(data)
        vs=sum(int(r.get('세액',0) or 0) for r in data if str(r.get('세무','')) in ['과세매입','카드매입','매입불공제'])
        if key=='불공제_조정필요': tv=vs
        fill=mf(cm[key]); extra=vs if key in ['불공제_조정필요','세무구분_오류'] else '-'
        for ci2,val in enumerate([sc['label'],cnt,extra],1):
            c=ws.cell(row=ri,column=ci2,value=val); c.fill=fill; c.border=mb(); c.alignment=Alignment(horizontal='center',vertical='center')
        if key in ['불공제_조정필요','세무구분_오류']: ws.cell(row=ri,column=3).number_format='#,##0'
    tr=6+len(SECTION_CONFIG); ws.cell(row=tr,column=1,value='신규 불공제 세액 합계').font=Font(bold=True)
    c=ws.cell(row=tr,column=3,value=tv); c.number_format='#,##0'; c.font=Font(bold=True,color='C00000')
    for ci in range(1,4): ws.cell(row=tr,column=ci).border=mb(); ws.cell(row=tr,column=ci).fill=mf(COLOR['summary_bg'])
    ws2=wb.create_sheet('검토결과'); ws2.freeze_panes='A2'
    for ci,col in enumerate([COL_HEADERS.get(c,c) for c in BASE_COLS+['_불공제사유','_비고','_오류유형']],1):
        ws2.column_dimensions[get_column_letter(ci)].width=COL_WIDTHS.get(col,12)
    cur=1
    for sk in SECTION_CONFIG: cur=write_section(ws2,results,sk,cur)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── UI 렌더링 ─────────────────────────────────────────

# 메인 헤더
st.markdown("""
<div class="main-header">
    <h1>🔍 ㈜비나우 매입세액불공제 자동 검출</h1>
    <p>ERP 매입 전표 데이터를 업로드하면 불공제 항목을 자동으로 검출하고 결과 엑셀 파일을 생성합니다.</p>
</div>
""", unsafe_allow_html=True)

# 탭 구성
tab1, tab2, tab3 = st.tabs(['📋 사용 가이드', '🔍 검토 실행', '⚙️ 설정'])

# ══════════════════════════════════════════════
# 탭1: 사용 가이드
# ══════════════════════════════════════════════
with tab1:
    st.markdown("### 📌 프로그램 개요")
    st.info("ERP에서 추출한 매입 전표 데이터(xlsx/csv)를 업로드하면, **과세매입·카드매입** 항목 중 불공제 대상을 자동으로 검출합니다.")

    st.markdown("---")
    st.markdown("### 🚀 사용 방법")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <div class="result-card card-blue" style="text-align:center; height:160px;">
            <div style="font-size:2rem;">📁</div>
            <div class="step-badge">1</div><strong>파일 업로드</strong>
            <p style="font-size:0.82rem; color:#555; margin-top:0.5rem;">ERP에서 추출한 매입 전표<br>xlsx 또는 csv 파일 업로드</p>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="result-card card-blue" style="text-align:center; height:160px;">
            <div style="font-size:2rem;">⚙️</div>
            <div class="step-badge">2</div><strong>검토 실행</strong>
            <p style="font-size:0.82rem; color:#555; margin-top:0.5rem;">검토 시작 버튼 클릭<br>자동으로 전표 분석</p>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div class="result-card card-blue" style="text-align:center; height:160px;">
            <div style="font-size:2rem;">📥</div>
            <div class="step-badge">3</div><strong>결과 다운로드</strong>
            <p style="font-size:0.82rem; color:#555; margin-top:0.5rem;">결과 엑셀 파일 다운로드<br>ERP에 수기 반영</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📊 결과 파일 구성")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
        <div class="guide-box">
            <h4>🔴 불공제 조정 필요</h4>
            <p>ERP에서 <strong>매입불공제</strong>로 세무구분을 변경해야 하는 항목입니다.<br>▶ 불공제사유(ERP입력용) 컬럼을 참고해서 사유도 함께 입력하세요.</p>
        </div>
        <div class="guide-box">
            <h4>⚠️ 세무구분 오류</h4>
            <p>세액=0인데 과세매입으로 분류되거나, 세액이 있는데 영세/면세매입으로 분류된 오류 항목입니다.</p>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown("""
        <div class="guide-box">
            <h4>🟡 담당자 확인 필요</h4>
            <p>적요만으로 공제/불공제 판단이 어려운 항목입니다.<br>비고란을 참고해서 담당자가 직접 확인 후 처리하세요.</p>
        </div>
        <div class="guide-box">
            <h4>📋 적요 공란</h4>
            <p>적요가 비어있는 전표입니다.<br>내용을 확인하고 적요를 기재해 주세요.</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### ✅ 자동 검출 기준 요약")
    criteria = [
        ("🚕 택시·대리운전·여객항공", "불공제", "여객운송업자는 세금계산서 발급 불가"),
        ("🚗 법인차량 관련 비용", "불공제", "비영업용소형승용차 (레이 경차 제외)"),
        ("📮 우체국 (세액=0)", "불공제", "면세 거래 (직접방문 접수)"),
        ("🏠 지사 파견 직원 사택", "불공제", "사업과 관련없는 지출"),
        ("💰 수입관세", "불공제", "납부 추적 불가로 보수적 처리"),
        ("📄 이니시스 + 서류발급(CFS)", "불공제", "세금계산서 발급 불가 구조"),
        ("🍽️ 외부인 접대비", "불공제", "기업업무추진비 (접대비관련매입세액)"),
        ("🗾 일본·미국 매장/아지트 관련", "담당자 확인", "선급금 또는 비용 여부 확인 필요"),
        ("🛒 미국 아마존 관련", "담당자 확인", "본사/비나우뷰티 귀속 여부 확인 필요"),
        ("☕ 커피챗 (비용센터 불명)", "담당자 확인", "면접 vs 행사 여부 확인 필요"),
    ]
    for name, result, reason in criteria:
        color = "card-red" if result == "불공제" else "card-yellow"
        badge_color = "#C00000" if result == "불공제" else "#BF8F00"
        st.markdown(f"""
        <div class="result-card {color}" style="display:flex; align-items:center; gap:1rem; padding:0.7rem 1rem;">
            <div style="min-width:200px;"><strong>{name}</strong></div>
            <div style="min-width:100px;"><span style="background:{badge_color};color:white;padding:2px 10px;border-radius:20px;font-size:0.8rem;">{result}</span></div>
            <div style="color:#555; font-size:0.85rem;">{reason}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### ⚠️ 주의사항")
    st.warning("""
    - 이 프로그램은 **적요 키워드 기반** 검출이므로, 적요가 불명확하게 기재된 경우 누락될 수 있습니다.
    - **담당자 확인 필요** 항목은 반드시 수동으로 확인 후 처리해 주세요.
    - 검출 기준 변경이 필요한 경우 **⚙️ 설정** 탭에서 키워드를 수정할 수 있습니다.
    - 업로드된 파일은 검토 후 서버에 저장되지 않습니다.
    """)

# ══════════════════════════════════════════════
# 탭2: 검토 실행
# ══════════════════════════════════════════════
with tab2:
    st.markdown("### 📁 파일 업로드")
    uploaded = st.file_uploader(
        "ERP에서 추출한 매입 전표 파일을 업로드하세요 (xlsx, xls, csv)",
        type=['xlsx','xls','csv'],
        help="파일은 서버에 저장되지 않으며, 검토 후 즉시 삭제됩니다."
    )

    if uploaded:
        st.success(f"✅ 파일 업로드 완료: **{uploaded.name}**")

        if st.button('🔍 검토 시작', type='primary', use_container_width=True):
            with st.spinner('전표 데이터를 분석 중입니다...'):
                try:
                    df = load_data(uploaded)
                    cfg = st.session_state.config
                    results = run_check(df.copy(), cfg)
                    st.session_state.results = results
                    st.session_state.filename = uploaded.name
                    st.success('검토가 완료됐습니다!')
                except Exception as e:
                    st.error(f'오류가 발생했습니다: {e}')

    if 'results' in st.session_state:
        results = st.session_state.results
        st.markdown("---")
        st.markdown("### 📊 검토 결과")

        # 통계
        c1,c2,c3,c4 = st.columns(4)
        n = [len(results[k]) for k in ['불공제_조정필요','담당자_확인필요','세무구분_오류','적요_공란']]
        vat_sum = sum(int(r.get('세액',0) or 0) for r in results['불공제_조정필요'] if str(r.get('세무','')) in ['과세매입','카드매입'])
        with c1: st.markdown(f'<div class="stat-box"><div class="num num-red">{n[0]}</div><div class="lbl">🔴 불공제 조정 필요</div></div>', unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="stat-box"><div class="num num-yellow">{n[1]}</div><div class="lbl">🟡 담당자 확인 필요</div></div>', unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="stat-box"><div class="num num-blue">{n[2]}</div><div class="lbl">⚠️ 세무구분 오류</div></div>', unsafe_allow_html=True)
        with c4: st.markdown(f'<div class="stat-box"><div class="num num-gray">{n[3]}</div><div class="lbl">📋 적요 공란</div></div>', unsafe_allow_html=True)

        st.markdown(f"<br>💰 **신규 불공제 세액 합계: {vat_sum:,}원**", unsafe_allow_html=True)

        # 상세 결과
        st.markdown("---")
        section_labels = {'불공제_조정필요':'🔴 불공제 조정 필요','담당자_확인필요':'🟡 담당자 확인 필요','세무구분_오류':'⚠️ 세무구분 오류','적요_공란':'📋 적요 공란'}
        disp_cols = {
            '불공제_조정필요': ['(세금)계산서일','적요','세무','거래처','공급가액','세액','_불공제사유','_비고'],
            '담당자_확인필요': ['(세금)계산서일','적요','세무','거래처','공급가액','세액','_비고'],
            '세무구분_오류':   ['(세금)계산서일','적요','세무','거래처','공급가액','세액','_오류유형','_비고'],
            '적요_공란':       ['(세금)계산서일','세무','거래처','공급가액','세액'],
        }
        for key, label in section_labels.items():
            data = results[key]
            with st.expander(f"{label}  ({len(data)}건)", expanded=(len(data)>0 and key=='불공제_조정필요')):
                if not data:
                    st.info("해당 없음")
                else:
                    df_show = pd.DataFrame(data)
                    cols = [c for c in disp_cols[key] if c in df_show.columns]
                    rename = {'(세금)계산서일':'계산서일','_불공제사유':'불공제사유(ERP입력용)','_비고':'비고','_오류유형':'오류유형'}
                    st.dataframe(df_show[cols].rename(columns=rename), use_container_width=True, height=300)

        # 다운로드
        st.markdown("---")
        st.markdown("### 📥 결과 파일 다운로드")
        buf = make_excel(results, st.session_state.filename)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        st.download_button(
            label='📥 검토결과 엑셀 다운로드',
            data=buf,
            file_name=f'검토결과_{ts}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
            type='primary',
        )
    else:
        st.info("📂 파일을 업로드하고 검토 시작 버튼을 눌러주세요.")

# ══════════════════════════════════════════════
# 탭3: 설정
# ══════════════════════════════════════════════
with tab3:
    st.markdown("### ⚙️ 검출 키워드 설정")
    st.info("키워드는 항목당 한 줄씩 입력하세요. 정규식 사용 가능합니다. (예: `카카오T.*택시`)")

    cfg = st.session_state.config

    with st.expander("🔴 불공제 키워드", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            taxi = st.text_area("택시·대리운전·항공비 키워드", value='\n'.join(cfg['taxi_keywords']), height=200, key='taxi')
        with c2:
            entertain = st.text_area("접대비 키워드", value='\n'.join(cfg['entertain_keywords']), height=200, key='entertain')

    with st.expander("🟡 담당자 확인 키워드", expanded=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            store = st.text_area("해외 매장·아지트 키워드", value='\n'.join(cfg['store_keywords']), height=150, key='store')
        with c2:
            amazon = st.text_area("미국 아마존 키워드", value='\n'.join(cfg['us_amazon_keywords']), height=150, key='amazon')
        with c3:
            exclude = st.text_area("담당자확인 제외 키워드", value='\n'.join(cfg['store_exclude_keywords']), height=150, key='exclude')

    with st.expander("✅ 공제 예외 차량번호"):
        cars = st.text_area("경차 차량번호 (공제 대상)", value='\n'.join(cfg['exempt_cars']), height=80, key='cars')

    with st.expander("📝 비고 문구"):
        note_g = st.text_area("일반 담당자확인 비고", value=cfg['note_general'], height=80, key='note_g')
        note_a = st.text_area("미국 아마존 담당자확인 비고", value=cfg['note_us_amazon'], height=80, key='note_a')

    if st.button('💾 설정 저장', type='primary', use_container_width=True):
        st.session_state.config = {
            'taxi_keywords': [l.strip() for l in taxi.splitlines() if l.strip()],
            'entertain_keywords': [l.strip() for l in entertain.splitlines() if l.strip()],
            'store_keywords': [l.strip() for l in store.splitlines() if l.strip()],
            'us_amazon_keywords': [l.strip() for l in amazon.splitlines() if l.strip()],
            'store_exclude_keywords': [l.strip() for l in exclude.splitlines() if l.strip()],
            'exempt_cars': [l.strip() for l in cars.splitlines() if l.strip()],
            'note_general': note_g.strip(),
            'note_us_amazon': note_a.strip(),
        }
        st.success('✅ 설정이 저장됐습니다! 다음 검토부터 반영됩니다.')
